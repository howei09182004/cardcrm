import base64
import json
import os
from datetime import datetime
from pathlib import Path

# Load .env file if present
_env_file = Path(__file__).parent / ".env"
if _env_file.exists():
    for _line in _env_file.read_text().splitlines():
        if "=" in _line and not _line.startswith("#"):
            _k, _v = _line.split("=", 1)
            os.environ.setdefault(_k.strip(), _v.strip())

import cv2
import numpy as np
from google import genai
from google.genai import types
import openpyxl
from flask import Flask, jsonify, render_template, request, send_file
from flask_cors import CORS

app = Flask(__name__)
CORS(app)

CONTACTS_FILE = Path(__file__).parent / "contacts.xlsx"
CARDS_DIR = Path(__file__).parent / "cards"
CARDS_DIR.mkdir(exist_ok=True)
HEADERS = [
    "name_kr", "name_en", "title_kr", "title_en",
    "company_kr", "company_en", "email",
    "phone_mobile", "phone_office", "phone_fax",
    "address", "website", "instagram", "notes", "save_type", "card_image"
]

CROP_ONLY_PROMPT = """Find the business card in this photo.
Return ONLY this JSON with the bounding box as fractions 0.0-1.0:
{"x1": 0.0, "y1": 0.0, "x2": 1.0, "y2": 1.0}
x1,y1=top-left, x2,y2=bottom-right. Be tight — no background. Return ONLY the JSON."""


def crop_only(image_data: bytes) -> bytes:
    """Use a dedicated Gemini call just for finding the card boundary, then crop."""
    import time
    import concurrent.futures
    client = genai.Client(api_key=os.environ.get("GEMINI_API_KEY", ""))
    last_err = None
    for attempt in range(4):
        try:
            with concurrent.futures.ThreadPoolExecutor(max_workers=1) as ex:
                future = ex.submit(
                    client.models.generate_content,
                    model="gemini-flash-lite-latest",
                    contents=[
                        types.Part.from_bytes(data=image_data, mime_type="image/jpeg"),
                        CROP_ONLY_PROMPT
                    ]
                )
                response = future.result(timeout=30)
            break
        except concurrent.futures.TimeoutError:
            last_err = Exception("Gemini timeout (30s)")
            wait = (attempt + 1) * 15
            print(f"[CROP] timeout, retry {attempt+1} after {wait}s", flush=True)
            time.sleep(wait)
        except Exception as e:
            last_err = e
            msg = str(e)
            if "503" in msg or "UNAVAILABLE" in msg or "429" in msg or "RESOURCE_EXHAUSTED" in msg:
                print(f"[CROP] 429/503 attempt {attempt+1}: {msg[:60]}", flush=True)
                raise  # 直接報錯，不 retry，讓使用者知道配額用完
            else:
                raise
    else:
        raise last_err
    text = response.text.strip()
    if text.startswith("```"):
        lines = text.split("\n")
        text = "\n".join(lines[1:] if lines[0].startswith("```") else lines)
    if text.endswith("```"):
        text = "\n".join(text.split("\n")[:-1])
    start = text.find("{")
    end = text.rfind("}") + 1
    data = json.loads(text[start:end])
    print(f"[CROP] raw: {data}", flush=True)
    if "box_2d" in data:
        b = data["box_2d"]
        if isinstance(b, list) and len(b) == 4:
            # [ymin, xmin, ymax, xmax] in 0-1000 scale
            bbox = {"x1": b[1]/1000, "y1": b[0]/1000, "x2": b[3]/1000, "y2": b[2]/1000}
        else:
            bbox = {"x1": b.get("xmin",0)/1000, "y1": b.get("ymin",0)/1000,
                    "x2": b.get("xmax",1000)/1000, "y2": b.get("ymax",1000)/1000}
    else:
        bbox = {"x1": data.get("x1", 0), "y1": data.get("y1", 0),
                "x2": data.get("x2", 1), "y2": data.get("y2", 1)}
    print(f"[CROP] bbox: {bbox}", flush=True)
    return grabcut_crop(image_data, bbox)


EXTRACTION_PROMPT = """This photo contains one or more business cards. Find the largest, most complete business card and extract its contact information.

CRITICAL RULES:
- name_en and name_kr must be a PERSON'S name only. Never put a building name, room number, floor number, company name, or address in the name field. If no person name is visible, use empty string.
- email must be a valid email address (contains @). Never put a URL, address, or other text in the email field.
- company_en / company_kr is the organization name, not a building or address.
- Do not put address text into any field except "address".
- The card may contain Korean, English, or Chinese text.

Return ONLY a JSON object with these exact fields:
{
  "bbox": {"x1": 0.0, "y1": 0.0, "x2": 1.0, "y2": 1.0},
  "name_en": "Person's English or romanized name only (e.g. John Kim)",
  "name_kr": "Person's Korean name in 한글 only (e.g. 김철수)",
  "name_cn": "Person's Chinese name only if present",
  "title_en": "Job title in English",
  "title_kr": "Job title in Korean if present",
  "company_en": "Company or organization name in English",
  "company_kr": "Company or organization name in Korean if present",
  "email": "email@domain.com (must contain @, else empty string)",
  "phone_mobile": "Mobile number (M / 010-...)",
  "phone_office": "Office phone (T / 02-...)",
  "phone_fax": "Fax number (F) if present",
  "address": "Full street/building address",
  "website": "Website URL",
  "instagram": "Instagram handle if present",
  "notes": "Any other info not fitting above fields"
}

For bbox: x1,y1 is top-left corner, x2,y2 is bottom-right corner, ALL values as fractions 0.0–1.0. Do NOT use pixel values.
Use empty string for any missing field. Return only the JSON, no other text."""


def get_or_create_workbook():
    if CONTACTS_FILE.exists():
        wb = openpyxl.load_workbook(CONTACTS_FILE)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Contacts"
        ws.append(HEADERS)
        for cell in ws[1]:
            cell.font = openpyxl.styles.Font(bold=True)
    return wb, ws


def save_to_excel(contact: dict) -> int:
    wb, ws = get_or_create_workbook()
    row = [contact.get(h, "") for h in HEADERS]
    ws.append(row)
    wb.save(CONTACTS_FILE)
    return ws.max_row - 1



def order_points(pts):
    rect = np.zeros((4, 2), dtype=np.float32)
    s = pts.sum(axis=1)
    rect[0] = pts[np.argmin(s)]
    rect[2] = pts[np.argmax(s)]
    diff = np.diff(pts, axis=1)
    rect[1] = pts[np.argmin(diff)]
    rect[3] = pts[np.argmax(diff)]
    return rect


def crop_card(image_data: bytes) -> bytes:
    arr = np.frombuffer(image_data, np.uint8)
    img = cv2.imdecode(arr, cv2.IMREAD_COLOR)
    if img is None:
        return image_data
    h, w = img.shape[:2]
    img_area = h * w

    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    best_rect = None
    best_score = 0

    for blur_size in [5, 11, 21]:
        blurred = cv2.GaussianBlur(gray, (blur_size, blur_size), 0)
        for thresh_method in ['canny', 'otsu', 'adaptive']:
            if thresh_method == 'canny':
                edges = cv2.Canny(blurred, 20, 80)
            elif thresh_method == 'otsu':
                _, edges = cv2.threshold(blurred, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
                edges = cv2.Canny(edges, 20, 80)
            else:
                edges = cv2.adaptiveThreshold(blurred, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
                                              cv2.THRESH_BINARY, 11, 2)
                edges = cv2.Canny(edges, 20, 80)
            kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (5, 5))
            edges = cv2.dilate(edges, kernel, iterations=2)
            edges = cv2.erode(edges, kernel, iterations=1)
            contours, _ = cv2.findContours(edges, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
            for cnt in contours:
                area = cv2.contourArea(cnt)
                if area < img_area * 0.05 or area > img_area * 0.92:
                    continue
                # Use minimum area rectangle for tight fit
                rot_rect = cv2.minAreaRect(cnt)
                rw, rh = rot_rect[1]
                if rw < 10 or rh < 10:
                    continue
                ratio = max(rw, rh) / max(min(rw, rh), 1)
                # Business card aspect ratio: 1.3 to 3.0 (allow square-ish cards too)
                if 1.2 <= ratio <= 3.2:
                    score = area / img_area
                    if score > best_score:
                        best_score = score
                        best_rect = rot_rect

    if best_rect is not None:
        box = cv2.boxPoints(best_rect).astype(np.float32)
        # Perspective transform using the 4 corners
        rect = order_points(box)
        wA = np.linalg.norm(rect[2] - rect[3])
        wB = np.linalg.norm(rect[1] - rect[0])
        hA = np.linalg.norm(rect[1] - rect[2])
        hB = np.linalg.norm(rect[0] - rect[3])
        maxW, maxH = int(max(wA, wB)), int(max(hA, hB))
        if maxW >= 10 and maxH >= 10:
            dst = np.array([[0, 0], [maxW-1, 0], [maxW-1, maxH-1], [0, maxH-1]], dtype=np.float32)
            M = cv2.getPerspectiveTransform(rect, dst)
            cropped = cv2.warpPerspective(img, M, (maxW, maxH))
            if maxH > maxW:
                cropped = cv2.rotate(cropped, cv2.ROTATE_90_CLOCKWISE)
            _, buf = cv2.imencode('.jpg', cropped, [cv2.IMWRITE_JPEG_QUALITY, 95])
            return buf.tobytes()

    return image_data


def extract_from_image(image_data: bytes, media_type: str) -> dict:
    client = genai.Client(api_key=os.environ.get("GEMINI_API_KEY", ""))
    response = client.models.generate_content(
        model="gemini-flash-lite-latest",
        contents=[
            types.Part.from_bytes(data=image_data, mime_type=media_type),
            EXTRACTION_PROMPT
        ]
    )
    text = response.text.strip()
    start = text.find("{")
    end = text.rfind("}") + 1
    return json.loads(text[start:end])


def find_card_by_color(img: np.ndarray) -> np.ndarray | None:
    """Find the largest white/light rectangular region (the business card)."""
    h, w = img.shape[:2]
    hsv = cv2.cvtColor(img, cv2.COLOR_BGR2HSV)
    # Mask: low saturation (not colorful) + high brightness (light/white)
    mask = cv2.inRange(hsv, (0, 0, 160), (180, 60, 255))
    kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (20, 20))
    mask = cv2.morphologyEx(mask, cv2.MORPH_CLOSE, kernel)
    mask = cv2.morphologyEx(mask, cv2.MORPH_OPEN, kernel)
    contours, _ = cv2.findContours(mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    best = None
    best_score = 0
    for cnt in contours:
        area = cv2.contourArea(cnt)
        if area < w * h * 0.04:
            continue
        x, y, bw, bh = cv2.boundingRect(cnt)
        ratio = max(bw, bh) / max(min(bw, bh), 1)
        if ratio > 4:  # too elongated, not a card
            continue
        if area > best_score:
            best_score = area
            best = (x, y, bw, bh)
    if best is None:
        return None
    x, y, bw, bh = best
    pad = 10
    x1 = max(0, x - pad)
    y1 = max(0, y - pad)
    x2 = min(w, x + bw + pad)
    y2 = min(h, y + bh + pad)
    return img[y1:y2, x1:x2]


def grabcut_crop(image_data: bytes, bbox: dict) -> bytes:
    """Gemini bbox rough crop, then Canny edge detection to find tight card boundary."""
    arr = np.frombuffer(image_data, np.uint8)
    img = cv2.imdecode(arr, cv2.IMREAD_COLOR)
    if img is None:
        return image_data
    h, w = img.shape[:2]
    try:
        def to_frac(v, dim):
            return v / dim if v > 1 else v
        pad = 0.02
        x1 = max(0.0, to_frac(bbox["x1"], w) - pad)
        y1 = max(0.0, to_frac(bbox["y1"], h) - pad)
        x2 = min(1.0, to_frac(bbox["x2"], w) + pad)
        y2 = min(1.0, to_frac(bbox["y2"], h) + pad)
        rx1, ry1 = int(x1 * w), int(y1 * h)
        rx2, ry2 = int(x2 * w), int(y2 * h)
        if rx2 - rx1 < 50 or ry2 - ry1 < 50:
            return image_data
        rough = img[ry1:ry2, rx1:rx2].copy()
        rh, rw = rough.shape[:2]

        # Edge detection to find the card's sharp rectangular boundary
        gray = cv2.cvtColor(rough, cv2.COLOR_BGR2GRAY)
        blurred = cv2.GaussianBlur(gray, (5, 5), 0)
        edges = cv2.Canny(blurred, 15, 60)
        edges = cv2.dilate(edges, np.ones((3, 3), np.uint8), iterations=2)

        contours, _ = cv2.findContours(edges, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        # Only keep contours covering at least 15% of rough area
        min_area = rw * rh * 0.15
        valid = [c for c in contours if cv2.contourArea(c) > min_area]

        if valid:
            all_pts = np.vstack(valid)
            gx, gy, gw, gh = cv2.boundingRect(all_pts)
            # Only use if tighter than rough (avoids making things worse)
            if gw * gh < rw * rh * 0.95:
                pad_px = 4
                gx = max(0, gx - pad_px)
                gy = max(0, gy - pad_px)
                gw = min(rw - gx, gw + pad_px * 2)
                gh = min(rh - gy, gh + pad_px * 2)
                final = rough[gy:gy+gh, gx:gx+gw]
                if final.size > 0:
                    if final.shape[0] > final.shape[1]:
                        final = cv2.rotate(final, cv2.ROTATE_90_CLOCKWISE)
                    _, buf = cv2.imencode('.jpg', final, [cv2.IMWRITE_JPEG_QUALITY, 95])
                    return buf.tobytes()

        # Fallback: use rough crop directly
        final = rough
        if final.shape[0] > final.shape[1]:
            final = cv2.rotate(final, cv2.ROTATE_90_CLOCKWISE)
        _, buf = cv2.imencode('.jpg', final, [cv2.IMWRITE_JPEG_QUALITY, 95])
        return buf.tobytes()
    except Exception as e:
        print(f"[CROP] crop error: {e}", flush=True)
        return image_data


def is_background_crop(img_region: np.ndarray) -> bool:
    """Return True if the region is mostly background (wood/table) rather than a card."""
    hsv = cv2.cvtColor(img_region, cv2.COLOR_BGR2HSV)
    avg_h = float(hsv[:,:,0].mean())
    avg_s = float(hsv[:,:,1].mean())
    avg_v = float(hsv[:,:,2].mean())
    # Wood/table: brownish hue, moderate saturation, not too dark
    is_wood = (8 < avg_h < 32 and avg_s > 45 and avg_v > 80)
    # Also bad if the card occupies very little area relative to crop size
    # Check what fraction of the crop is "card-like" (not wood)
    wood_mask = cv2.inRange(hsv, (8, 45, 80), (32, 255, 255))
    wood_frac = float(wood_mask.sum() / 255) / (img_region.shape[0] * img_region.shape[1])
    return is_wood and wood_frac > 0.5


def tight_crop_within(img: np.ndarray) -> np.ndarray:
    """Try to find and tightly crop a card within an already-roughly-cropped image."""
    h, w = img.shape[:2]
    img_area = h * w
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    best_rect = None
    best_score = 0
    for blur_size in [3, 7, 15]:
        blurred = cv2.GaussianBlur(gray, (blur_size, blur_size), 0)
        for thresh in [cv2.Canny(blurred, 15, 60),
                       cv2.Canny(cv2.threshold(blurred, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)[1], 15, 60)]:
            kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (5, 5))
            edges = cv2.dilate(thresh, kernel, iterations=2)
            edges = cv2.erode(edges, kernel, iterations=1)
            contours, _ = cv2.findContours(edges, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
            for cnt in contours:
                area = cv2.contourArea(cnt)
                # Must be large portion of the rough crop (card should dominate)
                if area < img_area * 0.25 or area > img_area * 0.97:
                    continue
                rot_rect = cv2.minAreaRect(cnt)
                rw, rh = rot_rect[1]
                if rw < 20 or rh < 20:
                    continue
                ratio = max(rw, rh) / max(min(rw, rh), 1)
                if 1.1 <= ratio <= 3.5:
                    score = area / img_area
                    if score > best_score:
                        best_score = score
                        best_rect = rot_rect
    if best_rect is not None and best_score < 0.96:
        box = cv2.boxPoints(best_rect).astype(np.float32)
        rect = order_points(box)
        wA = np.linalg.norm(rect[2] - rect[3])
        wB = np.linalg.norm(rect[1] - rect[0])
        hA = np.linalg.norm(rect[1] - rect[2])
        hB = np.linalg.norm(rect[0] - rect[3])
        maxW, maxH = int(max(wA, wB)), int(max(hA, hB))
        if maxW >= 30 and maxH >= 30:
            dst = np.array([[0,0],[maxW-1,0],[maxW-1,maxH-1],[0,maxH-1]], dtype=np.float32)
            M = cv2.getPerspectiveTransform(rect, dst)
            warped = cv2.warpPerspective(img, M, (maxW, maxH))
            if maxH > maxW:
                warped = cv2.rotate(warped, cv2.ROTATE_90_CLOCKWISE)
            return warped
    return img


def gemini_bbox_crop(image_data: bytes, bbox: dict) -> bytes:
    """Crop image using Gemini-provided bounding box, then refine with edge detection."""
    arr = np.frombuffer(image_data, np.uint8)
    img = cv2.imdecode(arr, cv2.IMREAD_COLOR)
    if img is None:
        return image_data
    h, w = img.shape[:2]
    try:
        def to_frac(v, dim):
            return v / dim if v > 1 else v
        x1 = max(0, int(to_frac(bbox["x1"], w) * w))
        y1 = max(0, int(to_frac(bbox["y1"], h) * h))
        x2 = min(w, int(to_frac(bbox["x2"], w) * w))
        y2 = min(h, int(to_frac(bbox["y2"], h) * h))
        if x2 - x1 < 20 or y2 - y1 < 20:
            return image_data
        rough = img[y1:y2, x1:x2]
        if is_background_crop(rough):
            # Gemini missed — try OpenCV on full image
            opencv_result = crop_card(image_data)
            ocv_arr = np.frombuffer(opencv_result, np.uint8)
            ocv_img = cv2.imdecode(ocv_arr, cv2.IMREAD_COLOR)
            if ocv_img is not None and not is_background_crop(ocv_img):
                return opencv_result
            # Still bad — just use bbox anyway
        else:
            # Stage 2: refine within the rough crop
            refined = tight_crop_within(rough)
            rough = refined
        _, buf = cv2.imencode('.jpg', rough, [cv2.IMWRITE_JPEG_QUALITY, 95])
        return buf.tobytes()
    except Exception:
        return image_data


def gemini_perspective_crop(image_data: bytes, corners: dict) -> bytes:
    """Perspective-correct crop using Gemini-provided four corners."""
    arr = np.frombuffer(image_data, np.uint8)
    img = cv2.imdecode(arr, cv2.IMREAD_COLOR)
    if img is None:
        return image_data
    h, w = img.shape[:2]
    try:
        def to_px(val, dim):
            # Gemini sometimes returns fractions, sometimes pixel values
            return val if val > 1 else val * dim

        pts = np.array([
            [to_px(corners["tl"][0], w), to_px(corners["tl"][1], h)],
            [to_px(corners["tr"][0], w), to_px(corners["tr"][1], h)],
            [to_px(corners["br"][0], w), to_px(corners["br"][1], h)],
            [to_px(corners["bl"][0], w), to_px(corners["bl"][1], h)],
        ], dtype=np.float32)
        rect = order_points(pts)
        wA = np.linalg.norm(rect[2] - rect[3])
        wB = np.linalg.norm(rect[1] - rect[0])
        hA = np.linalg.norm(rect[1] - rect[2])
        hB = np.linalg.norm(rect[0] - rect[3])
        maxW, maxH = int(max(wA, wB)), int(max(hA, hB))
        if maxW < 20 or maxH < 20:
            return image_data
        dst = np.array([[0, 0], [maxW-1, 0], [maxW-1, maxH-1], [0, maxH-1]], dtype=np.float32)
        M = cv2.getPerspectiveTransform(rect, dst)
        warped = cv2.warpPerspective(img, M, (maxW, maxH))
        if maxH > maxW:
            warped = cv2.rotate(warped, cv2.ROTATE_90_CLOCKWISE)
        _, buf = cv2.imencode('.jpg', warped, [cv2.IMWRITE_JPEG_QUALITY, 95])
        return buf.tobytes()
    except Exception:
        return image_data


def load_all_contacts() -> list:
    if not CONTACTS_FILE.exists():
        return []
    wb = openpyxl.load_workbook(CONTACTS_FILE)
    ws = wb.active
    contacts = []
    headers = [cell.value for cell in ws[1]]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(row):
            contacts.append(dict(zip(headers, row)))
    return contacts


@app.route("/")
def index():
    return render_template("index.html")


@app.route("/cards/<filename>")
def serve_card(filename):
    return send_file(CARDS_DIR / filename, mimetype="image/jpeg")


@app.route("/api/scan", methods=["POST"])
def scan():
    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400
    file = request.files["file"]
    ext = file.filename.rsplit(".", 1)[-1].lower()
    media_type_map = {"jpg": "image/jpeg", "jpeg": "image/jpeg", "png": "image/png", "webp": "image/webp", "pdf": "application/pdf"}
    media_type = media_type_map.get(ext, "image/jpeg")
    try:
        image_data = file.read()
        contact = extract_from_image(image_data, media_type)
        bbox = contact.pop("bbox", None)
        if bbox and media_type != "application/pdf":
            image_data = gemini_bbox_crop(image_data, bbox)
        # Save card image
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        name = (contact.get("name_en") or contact.get("name_kr") or "card").replace(" ", "_")
        img_filename = f"{ts}_{name}.jpg"
        img_path = CARDS_DIR / img_filename
        img_path.write_bytes(image_data)
        contact["card_image"] = f"/cards/{img_filename}"
        contact["card_path"] = str(img_path)
        return jsonify({"success": True, "contact": contact})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/save", methods=["POST"])
def save():
    contact = request.json
    try:
        row = save_to_excel(contact)
        return jsonify({"success": True, "row": row})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/optin", methods=["POST"])
def optin():
    contact = request.json
    contact["save_type"] = "optin"
    try:
        row = save_to_excel(contact)
        # TODO: send to FluentCRM webhook with opt-in tag when webhook URL is configured
        return jsonify({"success": True, "row": row})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/contacts", methods=["GET"])
def contacts():
    return jsonify(load_all_contacts())


def normalize_phone(p: str) -> str:
    return "".join(c for c in (p or "") if c.isdigit())


def is_logo_only(c: dict) -> bool:
    """Card has no name, no email, no phone — likely a logo/company-only side."""
    key_fields = ["name_en", "name_kr", "email", "phone_mobile", "phone_office"]
    return not any((c.get(f) or "").strip() for f in key_fields)


def match_score(a: dict, b: dict) -> int:
    # Logo-only cards are never merged with anything
    if is_logo_only(a) or is_logo_only(b):
        return 0
    # Only merge on exact email match
    email_a = (a.get("email") or "").strip().lower()
    email_b = (b.get("email") or "").strip().lower()
    if email_a and email_b and email_a == email_b:
        return 100
    return 0


def merge_contacts(a: dict, b: dict) -> dict:
    merged = {}
    fields = ["name_kr", "name_en", "name_cn", "title_kr", "title_en",
              "company_kr", "company_en", "email", "phone_mobile",
              "phone_office", "phone_fax", "address", "website", "instagram", "notes"]
    for f in fields:
        va, vb = (a.get(f) or "").strip(), (b.get(f) or "").strip()
        if va and vb and va != vb:
            merged[f] = va if len(va) >= len(vb) else vb
        else:
            merged[f] = va or vb
    # Keep both card images
    imgs = [x for x in [a.get("card_image"), b.get("card_image")] if x]
    merged["card_image"] = imgs[0] if imgs else ""
    merged["card_image_back"] = imgs[1] if len(imgs) > 1 else ""
    paths = [x for x in [a.get("card_path"), b.get("card_path")] if x]
    merged["card_path"] = paths[0] if paths else ""
    merged["card_path_back"] = paths[1] if len(paths) > 1 else ""
    return merged


@app.route("/crop-test")
def crop_test_page():
    return render_template("crop_test.html")


@app.route("/api/crop-test", methods=["POST"])
def api_crop_test():
    if "file" not in request.files:
        return jsonify({"error": "No file"}), 400
    file = request.files["file"]
    try:
        image_data = file.read()
        cropped = crop_only(image_data)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        fname = f"croptest_{ts}.jpg"
        (CARDS_DIR / fname).write_bytes(cropped)
        return jsonify({"success": True, "url": f"/cards/{fname}"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/crop-batch", methods=["POST"])
def api_crop_batch():
    """Stream crop results one by one as SSE so the frontend updates in real time."""
    import time
    from flask import stream_with_context, Response as FlaskResponse
    files = request.files.getlist("files")
    if not files:
        return jsonify({"error": "No files"}), 400
    # Read all file data upfront (can't read after response starts)
    file_items = [(f.filename, f.read()) for f in files]

    def generate():
        for i, (fname, data) in enumerate(file_items):
            try:
                cropped = crop_only(data)
                ts = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
                out_fname = f"croptest_{ts}.jpg"
                (CARDS_DIR / out_fname).write_bytes(cropped)
                row = json.dumps({"i": i, "success": True, "url": f"/cards/{out_fname}", "name": fname})
            except Exception as e:
                row = json.dumps({"i": i, "success": False, "error": str(e), "name": fname})
            yield f"data: {row}\n\n"
        yield "data: {\"done\": true}\n\n"

    return FlaskResponse(
        stream_with_context(generate()),
        mimetype="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"}
    )


@app.route("/api/ocr-image", methods=["POST"])
def api_ocr_image():
    """OCR a single image (e.g. a stitched pair) and return structured contact data."""
    if "file" not in request.files:
        return jsonify({"error": "No file"}), 400
    file = request.files["file"]
    import time
    image_data = file.read()
    mime = file.content_type or "image/jpeg"
    last_err = None
    for attempt in range(4):
        try:
            data = extract_from_image(image_data, mime)
            return jsonify({"success": True, "data": data})
        except Exception as e:
            last_err = e
            msg = str(e)
            if "503" in msg or "429" in msg or "UNAVAILABLE" in msg or "RESOURCE_EXHAUSTED" in msg:
                wait = (attempt + 1) * 5
                time.sleep(wait)
            else:
                break
    return jsonify({"error": str(last_err)}), 500


@app.route("/api/scan-pair", methods=["POST"])
def scan_pair():
    """Stitch two card images side-by-side, then OCR as one combined image."""
    files = request.files.getlist("files")
    if len(files) != 2:
        return jsonify({"error": "Need exactly 2 files"}), 400
    try:
        imgs = []
        raw_datas = []
        for f in files:
            data = f.read()
            raw_datas.append(data)
            arr = np.frombuffer(data, np.uint8)
            img = cv2.imdecode(arr, cv2.IMREAD_COLOR)
            if img is None:
                return jsonify({"error": "Could not decode image"}), 400
            # Resize to same height for stitching
            imgs.append(img)

        # Resize both to same height
        h = min(imgs[0].shape[0], imgs[1].shape[0], 1200)
        resized = []
        for img in imgs:
            scale = h / img.shape[0]
            w = int(img.shape[1] * scale)
            resized.append(cv2.resize(img, (w, h)))

        # Stitch side by side with a small divider
        divider = np.ones((h, 20, 3), dtype=np.uint8) * 200
        combined = np.hstack([resized[0], divider, resized[1]])

        _, buf = cv2.imencode('.jpg', combined, [cv2.IMWRITE_JPEG_QUALITY, 92])
        combined_data = buf.tobytes()

        contact = extract_from_image(combined_data, "image/jpeg")
        bbox = contact.pop("bbox", None)

        # For paired scans, crop each original individually and keep both
        cropped_datas = []
        for data in raw_datas:
            arr = np.frombuffer(data, np.uint8)
            img = cv2.imdecode(arr, cv2.IMREAD_COLOR)
            result = crop_card(data)
            cropped_datas.append(result)

        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        name = (contact.get("name_en") or contact.get("name_kr") or "card").replace(" ", "_")

        img_filename = f"{ts}_{name}.jpg"
        img_path = CARDS_DIR / img_filename
        img_path.write_bytes(cropped_datas[0])

        img_filename_back = f"{ts}_{name}_back.jpg"
        img_path_back = CARDS_DIR / img_filename_back
        img_path_back.write_bytes(cropped_datas[1])

        contact["card_image"] = f"/cards/{img_filename}"
        contact["card_path"] = str(img_path)
        contact["card_image_back"] = f"/cards/{img_filename_back}"
        contact["card_path_back"] = str(img_path_back)

        return jsonify({"success": True, "contact": contact})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/merge", methods=["POST"])
def merge():
    contacts_list = request.json  # list of contact dicts
    used = set()
    result = []
    for i, c1 in enumerate(contacts_list):
        if i in used:
            continue
        best_j, best_score = -1, 0
        for j, c2 in enumerate(contacts_list):
            if j <= i or j in used:
                continue
            s = match_score(c1, c2)
            if s >= 80 and s > best_score:
                best_score = s
                best_j = j
        if best_j >= 0:
            used.add(i)
            used.add(best_j)
            merged = merge_contacts(c1, contacts_list[best_j])
            merged["_merged"] = True
            result.append(merged)
        else:
            result.append(c1)
    return jsonify(result)


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5002))
    app.run(debug=False, host="0.0.0.0", port=port)
